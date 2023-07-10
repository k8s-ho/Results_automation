import os
import openpyxl

# 입력 받을 디렉터리 경로
input_directory = input("입력 디렉터리 경로를 입력하세요: ")

# 결과 파일이 저장될 디렉터리 경로
output_directory = input("결과 디렉터리 경로를 입력하세요: ")

# 디렉터리 내의 파일 목록 얻기
file_list = os.listdir(input_directory)

for file_name in file_list:
    file_path = os.path.join(input_directory, file_name)
    if os.path.isfile(file_path):
        print(f"파일 처리 중: {file_name}")

        try:
            # 새 워크북 생성 및 활성 시트 선택
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # 컬럼 헤더 설정
            worksheet['A1'] = 'W-Code'
            worksheet['B1'] = '시스템 현황'

            # 데이터 입력을 시작할 행 번호 설정
            start_row = 2

            with open(file_path, 'r', encoding='utf-8') as file:
                code = file.read()

            # 코드에서 정보 추출
            lines = code.split('\n')

            u_code = ""
            system_status = ""

            for line in lines:
                if line.startswith('[U-'):
                    if u_code != "":
                        # 이전 항목의 정보를 저장
                        worksheet.cell(row=start_row, column=1, value=u_code)
                        worksheet.cell(row=start_row, column=2, value=system_status)
                        start_row += 1

                    # 새로운 U-Code 번호 추출
                    u_code = line.strip('[]')
                    system_status = ""
                elif line.startswith('·[COMPLETE]'):
                    continue
                elif line.strip() != "":
                    system_status += line.strip() + "\n"

            # 마지막 항목 정보 저장
            worksheet.cell(row=start_row, column=1, value=u_code)
            worksheet.cell(row=start_row, column=2, value=system_status)

            # U-Code 번호를 기준으로 정렬
            sorted_data = sorted(worksheet.iter_rows(min_row=2, values_only=True), key=lambda x: x[0])

            # 정렬된 데이터로 시트 업데이트
            sorted_worksheet = workbook.create_sheet('Sorted')
            sorted_worksheet['A1'] = 'U-Code'
            sorted_worksheet['B1'] = '시스템 현황'

            for row in sorted_data:
                sorted_worksheet.append(row)

            # 기존 시트 제거
            workbook.remove(workbook['Sheet'])

            # 결과 파일 경로 설정
            output_file_name = f"{os.path.splitext(file_name)[0]}_시스템_현황.xlsx"
            output_file_path = os.path.join(output_directory, output_file_name)

            # 워크북 저장
            workbook.save(output_file_path)

            print(f"파일 처리 완료: {file_name}")
            print(f"결과 파일 저장: {output_file_path}")
            print()
        except Exception as e:
            print(f"에러 발생 파일: {file_name}")
            print(f"에러 내용: {str(e)}")
            print()
